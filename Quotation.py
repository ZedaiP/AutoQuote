from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from copy import copy
import re
def normalize_text(text):
    if text is None:
        return ""
    # 全角转半角
    text = text.translate(str.maketrans(
        '　，。！？（）［］｛｝【】《》＂＇＾～｜＼',
        ' ,.!?()[]{}【】《》"\'^~|\\'
    ))
    # 去掉所有空白字符（包括空格、制表符、换行等）
    text = re.sub(r'\s+', '', text)
    return text

def find_b_column_header(ws, target_text):
    result = None
    target_text = normalize_text(target_text)
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if normalize_text(cell.value) == target_text:
            result = ws.cell(row=cell.row, column=3).value
            break
    # 4. 输出结果
    # print(f"找到 '{target_text}' 在B{cell.row}，对应C列值为: {result}")
    return result

if __name__ == "__main__":
    standard_font = Font(name="Calibri", size=12.5)
    standard_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


    result_path = 'LogicData/Format.xlsx'
    decorate_path = 'LogicData/Decorate.xlsx'

    Handling = 60
    UPUSDL = 34.3
    UPUSOH = 162.8

    # 打开xlsx文件
    wb_result = load_workbook(result_path)
    wb_decorate = load_workbook(decorate_path)
    wb_source = load_workbook('LogicData/Source.xlsx', data_only=True)

    # 示例：获取第一个工作表
    ws_result = wb_result.active
    ws_decorate = wb_decorate.active
    #获取样式
    styles = []
    for cell in ws_decorate[3]:
        styles.append({
            'font': copy(cell.font),  # 需要从copy模块导入
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        })
    # 遍历
    n = 3
    for sheet_name in wb_source.sheetnames:
        ws = wb_source[sheet_name]  # 按名称获取工作表
        print(f"正在处理工作表: {sheet_name}")
        ws_result.row_dimensions[n].height = 31
        for cell, style in zip(ws_result[n], styles):
            cell.font = style['font']
            cell.fill = style['fill']
            cell.border = style['border']
            cell.alignment = style['alignment']
            cell.number_format = style['number_format']

        B_labor_time = find_b_column_header(ws, "B    Labor time(s/pcs)") or 0
        T_labor_time = find_b_column_header(ws, "T     Labor time(s/pcs)") or 0
        Gold_labor_time = find_b_column_header(ws, "金 Labor time(s/pcs)") or 0
        total_time = B_labor_time + T_labor_time + Gold_labor_time
        ws_result[f'G{n}'] = f"={total_time}/0.8"

        D_labor_time = find_b_column_header(ws, "D    Labor time(s/pcs)") or 0
        ws_result[f'H{n}'] = f"={D_labor_time}/0.8"

        R_labor_time = find_b_column_header(ws, "R    Labor time(s/pcs)") or 0
        FiveD_labor_time = find_b_column_header(ws, "5D    Labor time(s/pcs)") or 0
        P_labor_time = find_b_column_header(ws, "P    Labor time(s/pcs)") or 0
        Routing_pressfit_working_time = R_labor_time + FiveD_labor_time + P_labor_time
        ws_result[f'I{n}'] = f"={Routing_pressfit_working_time}/0.8"

        SASM = find_b_column_header(ws, "SASM") or 0
        AVI = find_b_column_header(ws, "AVI") or 0
        INSP = find_b_column_header(ws, "INSP") or 0
        PACK = find_b_column_header(ws, "PACK") or 0
        Packing_working_time = SASM + AVI + INSP + PACK
        ws_result[f'J{n}'] = f"={Packing_working_time}/0.8"

        ICT = find_b_column_header(ws, "ICT") or 0
        ws_result[f'k{n}'] = ICT

        machine_time = find_b_column_header(ws, "Machine Time(s/pcs)") or 0
        ws_result[f'Q{n}'] = machine_time

        project_name = find_b_column_header(ws, "Project name") or ""
        ws_result[f'C{n}'] = project_name

        ws_result[f'N{n}'] = f"=SUM(G{n}:M{n})"
        ws_result[f'P{n}'] = f"=ROUND((O{n}*N{n}/3600)*IF(F{n}<1000,1.05,1.02),2)"
        ws_result[f'S{n}'] = f"=ROUND((R{n}*Q{n}/3600)*IF(F{n}<1000,1.05,1.02),2)"
        ws_result[f'T{n}'] = f"=SUM(P{n},S{n})"
        ws_result[f'U{n}'] = f"=F{n}"
        ws_result[f'W{n}'] = f"=V{n}*IF(F{n}<1000,0.004,0.002)"
        ws_result[f'X{n}'] = f"=ROUNDUP(P{n}*IF(F{n}<500,1.2,IF(F{n}<1000,1,IF(F{n}>5000,0.6,0.8))),2)"
        ws_result[f'Y{n}'] = f"=ROUNDUP(P{n}*IF(F{n}<500,1.5,IF(F{n}<1000,1.2,IF(F{n}>5000,0.6,0.9))),2)"
        ws_result[f'Z{n}'] = f"=ROUNDUP(SUM(T{n},W{n}:Y{n}),2)"

        ws_result[f'M{n}'] = Handling
        ws_result[f'O{n}'] = UPUSDL
        ws_result[f'R{n}'] = UPUSOH

        target_text = normalize_text("Panel Qty")
        for row in ws.iter_rows(min_col=3, max_col=3):
            cell = row[0]
            if normalize_text(cell.value) == target_text:
                pcs_panel = ws.cell(row=cell.row + 1, column=3).value
                break
        ws_result[f'E{n}'] = pcs_panel or 0
        print(f"写入E{n}，值：", pcs_panel)

        n += 1
    wb_result.save('output.xlsx')
