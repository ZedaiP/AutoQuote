import json
import os
import re
import sys
import xlrd
from copy import copy
from datetime import datetime

if getattr(sys, 'frozen', False):

    application_path = os.path.dirname(sys.executable)

    # 设置Qt插件路径
    os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = os.path.join(
        application_path, 'PyQt5', 'Qt5', 'plugins'
    )
    os.environ['QT_PLUGIN_PATH'] = os.path.join(
        application_path, 'PyQt5', 'Qt5', 'plugins'
    )

from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QFileDialog, QTextEdit, QProgressBar, QGroupBox,
                             QFormLayout, QMessageBox, QTabWidget)
from openpyxl import load_workbook


# =============================================================================
# 1. 核心逻辑工具函数 (保留原逻辑)
# =============================================================================

def normalize_text(text):
    if text is None:
        return ""
    text = str(text)
    # 全角转半角
    text = text.translate(str.maketrans(
        '　，。！？（）［］｛｝【】《》＂＇＾～｜＼',
        ' ,.!?()[]{}【】《》"\'^~|\\'
    ))
    # 去掉所有空白字符
    text = re.sub(r'\s+', '', text)
    return text


def find_b_column_header(ws, target_text):
    result = None
    target_text = normalize_text(target_text)
    for row in ws.iter_rows(min_col=2, max_col=2):
        try:
            cell = row[0]
            if normalize_text(cell.value) == target_text:
                result = ws.cell(row=cell.row, column=3).value
                break
        except:
            continue
    return result


def find_b_column_header_xlrd(ws, target_text):
    result = None
    target_text = normalize_text(target_text)

    for row_idx in range(ws.nrows):
        try:
            cell_value = ws.cell_value(row_idx, 1)

            if normalize_text(str(cell_value)) == target_text:
                result = ws.cell_value(row_idx, 2)
                break
        except:
            continue

    return result


def find_name_by_amount_xlrd(ws, target_row):
    if target_row < ws.nrows:
        try:
            cell_value = ws.cell_value(target_row, 3)
            if cell_value != 0:
                result = ws.cell_value(target_row, 1)
                amount = ws.cell_value(target_row, 2)
                return result, amount, cell_value
            else:
                return None, 0, 0  # 明确返回
        except Exception as e:
            print(f"读取第{target_row}行时出错: {e}")
            return None, 0, 0
    else:
        return None, 0, 0

    return None, 0, 0


# =============================================================================
# 2. 配置管理类
# =============================================================================
class ConfigManager:
    DEFAULT_CONFIG = {
        "paths": {
            "format_path": "LogicData/Format.xlsx",
            "decorate_path": "LogicData/Decorate.xlsx",
            "format_path_tab2": "LogicData/Format2.xlsx",
            "decorate_path_tab2": "LogicData/Decorate2.xlsx"
        },
        "params": {
            "Handling": 60,
            "UPUSDL": 34.3,
            "UPUSOH": 162.8,
            "StencilQty": 2,
            "SMTCarrierQty": 90,
            "Qty": 1
        }
    }
    CONFIG_FILE = "settings.json"

    @staticmethod
    def load_config():
        if not os.path.exists(ConfigManager.CONFIG_FILE):
            ConfigManager.save_config(ConfigManager.DEFAULT_CONFIG)
            return ConfigManager.DEFAULT_CONFIG
        try:
            with open(ConfigManager.CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
            # 确保所有必需的参数都存在，如果缺少则使用默认值
            if 'paths' not in config:
                config['paths'] = ConfigManager.DEFAULT_CONFIG['paths'].copy()
            else:
                for key, value in ConfigManager.DEFAULT_CONFIG['paths'].items():
                    if key not in config['paths']:
                        config['paths'][key] = value
                        
            if 'params' not in config:
                config['params'] = ConfigManager.DEFAULT_CONFIG['params'].copy()
            else:
                for key, value in ConfigManager.DEFAULT_CONFIG['params'].items():
                    if key not in config['params']:
                        config['params'][key] = value
                        
            return config
        except:
            return ConfigManager.DEFAULT_CONFIG

    @staticmethod
    def save_config(config_data):
        with open(ConfigManager.CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)


# =============================================================================
# 3. 工作线程 (处理Excel，防止界面卡死)
# =============================================================================
class ExcelWorker(QThread):
    log_signal = pyqtSignal(str)  # 发送日志
    progress_signal = pyqtSignal(int)  # 发送进度 0-100
    finished_signal = pyqtSignal(bool, str)  # 完成信号(是否成功, 消息)

    def __init__(self, source_path, config):
        super().__init__()
        self.source_path = source_path
        self.config = config

    def run(self):
        try:
            self.log_signal.emit("正在初始化...")

            # 读取配置
            format_path = self.config['paths']['format_path']
            decorate_path = self.config['paths']['decorate_path']

            val_handling = float(self.config['params']['Handling'])
            val_upusdl = float(self.config['params']['UPUSDL'])
            val_upusoh = float(self.config['params']['UPUSOH'])

            # 检查文件存在
            if not os.path.exists(format_path):
                raise FileNotFoundError(f"找不到格式文件: {format_path}")
            if not os.path.exists(decorate_path):
                raise FileNotFoundError(f"找不到修饰文件: {decorate_path}")

         
            self.log_signal.emit("正在加载工作簿 (这可能需要几秒钟)...")

            wb_result = load_workbook(format_path)
            wb_source = load_workbook(self.source_path, data_only=True)
            wb_decorate = load_workbook(decorate_path)

            ws_result = wb_result.active
            ws_decorate = wb_decorate.active

            # 获取样式
            self.log_signal.emit("正在读取样式模板...")
            styles = []
            # 注意：原代码假定取第3行的样式
            for cell in ws_decorate[3]:
                styles.append({
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })

            sheet_names = wb_source.sheetnames
            total_sheets = len(sheet_names)
            n = 3

            for idx, sheet_name in enumerate(sheet_names):
                self.log_signal.emit(f"处理工作表 [{idx + 1}/{total_sheets}]: {sheet_name}")

                ws = wb_source[sheet_name]
                ws_result.row_dimensions[n].height = 31

                # 应用样式
                for cell, style in zip(ws_result[n], styles):
                    cell.font = style['font']
                    cell.fill = style['fill']
                    cell.border = style['border']
                    cell.alignment = style['alignment']
                    cell.number_format = style['number_format']

                # 计算
                B_labor = find_b_column_header(ws, "B    Labor time(s/pcs)") or 0
                T_labor = find_b_column_header(ws, "T     Labor time(s/pcs)") or 0
                G_labor = find_b_column_header(ws, "金 Labor time(s/pcs)") or 0
                ws_result[f'G{n}'] = f"={(B_labor + T_labor + G_labor)}/0.8"

                D_labor = find_b_column_header(ws, "D    Labor time(s/pcs)") or 0
                ws_result[f'H{n}'] = f"={D_labor}/0.8"

                R_labor = find_b_column_header(ws, "R    Labor time(s/pcs)") or 0
                FiveD_labor = find_b_column_header(ws, "5D    Labor time(s/pcs)") or 0
                P_labor = find_b_column_header(ws, "P    Labor time(s/pcs)") or 0
                ws_result[f'I{n}'] = f"={(R_labor + FiveD_labor + P_labor)}/0.8"

                SASM = find_b_column_header(ws, "SASM") or 0
                AVI = find_b_column_header(ws, "AVI") or 0
                INSP = find_b_column_header(ws, "INSP") or 0
                PACK = find_b_column_header(ws, "PACK") or 0
                ws_result[f'J{n}'] = f"={(SASM + AVI + INSP + PACK)}/0.8"

                ICT = find_b_column_header(ws, "ICT") or 0
                ws_result[f'K{n}'] = ICT

                machine_time = find_b_column_header(ws, "Machine Time(s/pcs)") or 0
                ws_result[f'Q{n}'] = machine_time

                project_name = find_b_column_header(ws, "Project name") or ""
                ws_result[f'C{n}'] = project_name

                # 公式
                ws_result[f'N{n}'] = f"=SUM(G{n}:M{n})"
                ws_result[f'P{n}'] = f"=ROUND((O{n}*N{n}/3600)*IF(F{n}<1000,1.05,1.02),2)"
                ws_result[f'S{n}'] = f"=ROUND((R{n}*Q{n}/3600)*IF(F{n}<1000,1.05,1.02),2)"
                ws_result[f'T{n}'] = f"=SUM(P{n},S{n})"
                ws_result[f'U{n}'] = f"=F{n}"
                ws_result[f'W{n}'] = f"=V{n}*IF(F{n}<1000,0.004,0.002)"
                ws_result[f'X{n}'] = f"=ROUNDUP(P{n}*IF(F{n}<500,1.2,IF(F{n}<1000,1,IF(F{n}>5000,0.6,0.8))),2)"
                ws_result[f'Y{n}'] = f"=ROUNDUP(P{n}*IF(F{n}<500,1.5,IF(F{n}<1000,1.2,IF(F{n}>5000,0.6,0.9))),2)"
                ws_result[f'Z{n}'] = f"=ROUNDUP(SUM(T{n},W{n}:Y{n}),2)"

                # === 使用配置中的变量 ===
                ws_result[f'M{n}'] = val_handling  # 原代码是 60
                ws_result[f'O{n}'] = val_upusdl  # 原代码是 34.3
                ws_result[f'R{n}'] = val_upusoh  # 原代码是 162.8

                # 查找 Panel Qty
                pcs_panel = 0
                target_text = normalize_text("Panel Qty")
                for row in ws.iter_rows(min_col=3, max_col=3):
                    try:
                        cell = row[0]
                        if normalize_text(cell.value) == target_text:
                            pcs_panel = ws.cell(row=cell.row + 1, column=3).value
                            break
                    except:
                        continue
                ws_result[f'E{n}'] = pcs_panel or 0

                n += 1
                # 更新进度条
                self.progress_signal.emit(int(((idx + 1) / total_sheets) * 100))

            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(self.source_path)
            output_filename = f"Output_IE_{timestamp}.xlsx"
            output_path = os.path.join(output_dir, output_filename)

            self.log_signal.emit(f"正在保存文件: {output_path}")
            wb_result.save(output_path)

            self.finished_signal.emit(True, f"处理完成！\n文件已保存至:\n{output_path}")

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            self.finished_signal.emit(False, f"发生错误: {str(e)}\n{error_msg}")


# =============================================================================
# 3.2 工作线程 (处理第二个选项卡的Excel)
# =============================================================================
class ExcelWorker2(QThread):
    log_signal = pyqtSignal(str)  # 发送日志
    progress_signal = pyqtSignal(int)  # 发送进度 0-100
    finished_signal = pyqtSignal(bool, str)  # 完成信号(是否成功, 消息)

    def __init__(self, source_path, config):
        super().__init__()
        self.source_path = source_path
        self.config = config

    def run(self):
        try:
            self.log_signal.emit("正在初始化...")

            # 读取配置
            format_path = self.config['paths']['format_path_tab2']
            decorate_path = self.config['paths']['decorate_path_tab2']

            val_stencilqty = float(self.config['params']['StencilQty'])
            val_smtqty = float(self.config['params']['SMTCarrierQty'])
            val_qty = float(self.config['params']['Qty'])

            # 检查文件存在
            if not os.path.exists(format_path):
                raise FileNotFoundError(f"找不到格式文件: {format_path}")
            if not os.path.exists(decorate_path):
                raise FileNotFoundError(f"找不到修饰文件: {decorate_path}")

            self.log_signal.emit("正在加载工作簿 (这可能需要几秒钟)...")

            wb_result = load_workbook(format_path)
            wb_decorate = load_workbook(decorate_path)

            ws_result = wb_result.active
            ws_decorate = wb_decorate.active

            #获取样式
            self.log_signal.emit("正在读取样式模板...")
            styles1 = []
            styles2 = []
            styles3 = []
            styles4 = []
            # 注意：原代码假定取第3行的样式
            for cell in ws_decorate[3]:
                styles1.append({
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })

            for cell in ws_decorate[4]:
                styles2.append({
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })

            for cell in ws_decorate[5]:
                styles3.append({
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })

            for cell in ws_decorate[6]:
                styles4.append({
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })

            n = 3
            m = 11

            # 检查是文件还是文件夹
            if os.path.isdir(self.source_path):
                # 遍历文件夹中的所有xlsx文件
                self.log_signal.emit(f"正在扫描文件夹: {self.source_path}")
                xlsx_files = [f for f in os.listdir(self.source_path) if f.lower().endswith('.xls')]
                if not xlsx_files:
                    raise FileNotFoundError(f"在文件夹 {self.source_path} 中未找到xls文件")

                self.log_signal.emit(f"找到 {len(xlsx_files)} 个xls文件")

                xlsx_files.sort(key=lambda x: [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', x)])

                total_files = len(xlsx_files)
                for file_idx, xlsx_file in enumerate(xlsx_files):

                    process_file_path = os.path.join(self.source_path, xlsx_file)
                    wb_source = xlrd.open_workbook(process_file_path)
                    ws = wb_source.sheet_by_index(0)
                    self.log_signal.emit(f"处理文件: {xlsx_file}")
                    
                    # 更新进度条
                    self.progress_signal.emit(int(((file_idx + 1) / total_files) * 100))

                    ws_result.insert_rows(n, 1)
                    ws_result.row_dimensions[n].height = 25
                    # 应用样式
                    for cell, style in zip(ws_result[n], styles1):
                        cell.font = style['font']
                        cell.fill = style['fill']
                        cell.border = style['border']
                        cell.alignment = style['alignment']
                        cell.number_format = style['number_format']

                    name_without_ext = xlsx_file.rsplit('.', 1)[0]
                    match = re.search(r'\d+\.(.+)', name_without_ext)
                    if match:
                        name = match.group(1).replace('_', '*')
                    else:
                        name = ""
                    ws_result[f'C{n}'] = name

                    st = find_b_column_header_xlrd(ws, "Stencil (Top  side)") or 0
                    sb = find_b_column_header_xlrd(ws, "Stencil (bottom side)") or 0
                    StencilUP = ((st + sb) / 2) * 1.1
                    ws_result[f'H{n}'] = StencilUP

                    SMTCarrierUP = (find_b_column_header_xlrd(ws, "SMT carrier") or 0) * 1.1
                    ws_result[f'K{n}'] = SMTCarrierUP

                    name, amount, qty = find_name_by_amount_xlrd(ws, 11)
                    if name != None:
                        ws_result[f'M{n}'] = val_qty
                        ws_result[f'N{n}'] = amount * 1.1
                        ws_result[f'O{n}'] = f"=M{n}*N{n}"
                    else:
                        ws_result.merge_cells(f'M{n}:O{n}')
                        ws_result[f'M{n}'] = "--"
                    ws_result[f'G{n}'] = val_stencilqty
                    ws_result[f'J{n}'] = val_smtqty
                    ws_result[f'Q{n}'] = 1
                    ws_result[f'B{n}'] = f"{n - 2}"



                    ws_result[f'I{n}'] = f"=G{n}*H{n}"
                    ws_result[f'L{n}'] = f"=J{n}*K{n}"
                    ws_result[f'S{n}'] = f"=Q{n}*R{n}"
                    ws_result[f'W{n}'] = f"=U{n}*V{n}"
                    ws_result[f'X{n}'] = f"=SUM(I{n}, L{n}, O{n}, P{n}, S{n}, W{n})"
                    ws_result[f'Y{n}'] = f"=X{n}*1.13"
                    #ws_result[f'H{n}'] = f"={(SASM + AVI + INSP + PACK)}/0.8"
                    n += 1
                    m += 1
                    for i in range(4,11):
                        name, amount, qty = find_name_by_amount_xlrd(ws, i)
                        if name != None:
                            # 应用样式
                            for cell, style in zip(ws_result[m], styles3):
                                cell.font = style['font']
                                cell.fill = style['fill']
                                cell.border = style['border']
                                cell.alignment = style['alignment']
                                cell.number_format = style['number_format']
                            ws_result[f'C{m}'] = name
                            ws_result[f'D{m}'] = qty
                            ws_result[f'E{m}'] = amount * 1.1
                            ws_result[f'G{m}'] = f"{n - 3}"
                            m += 1


                ws_result.insert_rows(n, 1)
                ws_result.row_dimensions[n].height = 24
                m += 1

                # 应用样式
                for cell, style in zip(ws_result[n], styles2):
                    cell.font = style['font']
                    cell.fill = style['fill']
                    cell.border = style['border']
                    cell.alignment = style['alignment']
                    cell.number_format = style['number_format']

                # 应用样式
                for cell, style in zip(ws_result[m], styles4):
                    cell.font = style['font']
                    cell.fill = style['fill']
                    cell.border = style['border']
                    cell.alignment = style['alignment']
                    cell.number_format = style['number_format']

                ws_result[f'I{n}'] = f"=SUM(I3:I{n - 1})"
                ws_result[f'L{n}'] = f"=SUM(L3:L{n - 1})"
                ws_result[f'P{n}'] = f"=SUM(P3:P{n - 1})"
                ws_result[f'S{n}'] = f"=SUM(S3:S{n - 1})"
                ws_result[f'W{n}'] = f"=SUM(W3:W{n - 1})"
                ws_result[f'X{n}'] = f"=SUM(X3:X{n - 1})"
                ws_result[f'Y{n}'] = f"=SUM(Y3:Y{n - 1})"



                for i in range(3,n):
                    ws_result[f'P{i}'] = f"=SUMIFS($F${n + 9}:$F${m - 1},$G${n + 9}:$G${m - 1},B{i})"

                for i in range(n + 9, m):
                    ws_result[f'F{i}'] = f"=D{i}*E{i}"

                for i in range(1, n-2):
                    for cell, style in zip(ws_result[m], styles3):
                        cell.font = style['font']
                        cell.fill = style['fill']
                        cell.border = style['border']
                        cell.alignment = style['alignment']
                        cell.number_format = style['number_format']
                    ws_result[f'C{m}'] = "AVI Carrier"
                    ws_result[f'D{m}'] = 4
                    ws_result[f'E{m}'] = 892
                    ws_result[f'F{m}'] = f"=D{m}*E{m}"
                    ws_result[f'G{m}'] = i
                    m += 1
                ws_result[f'F{m}'] = f"=SUM(F{n + 9}:F{m - 1})"
            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(self.source_path)
            output_filename = f"Output_ME_{timestamp}.xlsx"  # 使用不同的文件名前缀
            output_path = os.path.join(output_dir, output_filename)

            self.log_signal.emit(f"正在保存文件: {output_path}")
            wb_result.save(output_path)

            self.finished_signal.emit(True, f"处理完成！\n文件已保存至:\n{output_path}")

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            self.finished_signal.emit(False, f"发生错误: {str(e)}\n{error_msg}")


# =============================================================================
# 4. 主界面
# =============================================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_data = ConfigManager.load_config()
        self.init_ui()
        self.apply_stylesheet()

    def init_ui(self):
        self.setWindowTitle("Automated Quotation System")
        self.resize(900, 700)

        # 主部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 创建选项卡控件
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # IE选项卡
        self.tab1 = QWidget()
        self.tab_widget.addTab(self.tab1, "报价-IE")
        self.init_tab1()
        
        # ME选项卡
        self.tab2 = QWidget()
        self.tab_widget.addTab(self.tab2, "报价-ME")
        self.init_tab2()
        
        # 底部状态栏
        self.status_label = QLabel("就绪")
        self.statusBar().addWidget(self.status_label)
    
    def init_tab1(self):
        tab1_layout = QVBoxLayout(self.tab1)
        tab1_layout.setSpacing(15)
        tab1_layout.setContentsMargins(10, 10, 10, 10)

        # 1. 标题区
        title_label = QLabel("自动报价系统")
        title_label.setObjectName("TitleLabel")
        title_label.setAlignment(Qt.AlignCenter)
        tab1_layout.addWidget(title_label)

        # 2. 配置区域 (分组框)
        config_group = QGroupBox("参数设定")
        config_layout = QFormLayout()
        config_layout.setSpacing(10)

        # 路径选择辅助函数
        def create_file_picker(key, label_text):
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.setContentsMargins(0, 0, 0, 0)

            line_edit = QLineEdit()
            line_edit.setText(self.config_data['paths'].get(key, ""))
            line_edit.setReadOnly(False)  # 允许手动粘贴

            btn = QPushButton("浏览...")
            btn.setFixedWidth(80)
            btn.clicked.connect(lambda: self.browse_file(line_edit, key))

            layout.addWidget(line_edit)
            layout.addWidget(btn)
            config_layout.addRow(label_text, container)
            return line_edit

        self.input_format_path = create_file_picker('format_path', "格式模板:")
        self.input_decorate_path = create_file_picker('decorate_path', "样式模板:")

        # 参数配置
        param_layout = QHBoxLayout()

        self.input_handling = QLineEdit(str(self.config_data['params']['Handling']))
        self.input_upusdl = QLineEdit(str(self.config_data['params']['UPUSDL']))
        self.input_upusoh = QLineEdit(str(self.config_data['params']['UPUSOH']))

        # 验证器：只允许输入数字
        for inp in [self.input_handling, self.input_upusdl, self.input_upusoh]:
            inp.setFixedWidth(100)
            inp.setPlaceholderText("数字")

        param_layout.addWidget(QLabel("Handling:"))
        param_layout.addWidget(self.input_handling)
        param_layout.addSpacing(20)
        param_layout.addWidget(QLabel("U/P(US$)-DL(hr):"))
        param_layout.addWidget(self.input_upusdl)
        param_layout.addSpacing(20)
        param_layout.addWidget(QLabel("U/P(US$)-OH(hr):"))
        param_layout.addWidget(self.input_upusoh)
        param_layout.addStretch()

        config_layout.addRow("参数设置:", param_layout)

        # 保存配置按钮
        save_cfg_btn = QPushButton("保存")
        save_cfg_btn.setFixedWidth(90)
        save_cfg_btn.clicked.connect(self.save_current_config)
        config_layout.addRow("", save_cfg_btn)

        config_group.setLayout(config_layout)
        tab1_layout.addWidget(config_group)

        # 3. 操作区域
        action_group = QGroupBox("一键报价")
        action_layout = QVBoxLayout()

        # 源文件选择
        source_layout = QHBoxLayout()
        self.source_file_edit = QLineEdit()
        self.source_file_edit.setPlaceholderText("请选择要处理的xlsx文件...")
        self.source_file_edit.setReadOnly(True)

        select_source_btn = QPushButton("选择源文件")
        select_source_btn.setObjectName("PrimaryButton")
        select_source_btn.setCursor(Qt.PointingHandCursor)
        select_source_btn.clicked.connect(self.select_source_file)

        source_layout.addWidget(self.source_file_edit)
        source_layout.addWidget(select_source_btn)

        # 运行按钮
        self.run_btn = QPushButton("开始处理")
        self.run_btn.setObjectName("ActionButton")
        self.run_btn.setCursor(Qt.PointingHandCursor)
        self.run_btn.setFixedHeight(50)
        self.run_btn.setEnabled(False)  # 选择文件后启用
        self.run_btn.clicked.connect(self.start_processing)

        action_layout.addLayout(source_layout)
        action_layout.addSpacing(10)
        action_layout.addWidget(self.run_btn)
        action_group.setLayout(action_layout)
        tab1_layout.addWidget(action_group)

        # 4. 日志与进度
        log_group = QGroupBox("运行日志 (Logs)")
        log_layout = QVBoxLayout()

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setObjectName("LogArea")

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignCenter)

        log_layout.addWidget(self.log_area)
        log_layout.addWidget(self.progress_bar)
        log_group.setLayout(log_layout)

        tab1_layout.addWidget(log_group, stretch=1)
    
    def init_tab2(self):
        tab2_layout = QVBoxLayout(self.tab2)
        tab2_layout.setSpacing(15)
        tab2_layout.setContentsMargins(10, 10, 10, 10)
        
        # 第二个选项卡的内容
        title_label = QLabel("自动报价系统")
        title_label.setObjectName("TitleLabel")
        title_label.setAlignment(Qt.AlignCenter)
        tab2_layout.addWidget(title_label)

        # 2. 配置区域 (分组框)
        config_group = QGroupBox("参数设定")
        config_layout = QFormLayout()
        config_layout.setSpacing(10)

        # 路径选择辅助函数
        def create_file_picker(key, label_text):
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.setContentsMargins(0, 0, 0, 0)

            line_edit = QLineEdit()
            line_edit.setText(self.config_data['paths'].get(key, ""))
            line_edit.setReadOnly(False)  # 允许手动粘贴

            btn = QPushButton("浏览...")
            btn.setFixedWidth(80)
            btn.clicked.connect(lambda: self.browse_file(line_edit, key))

            layout.addWidget(line_edit)
            layout.addWidget(btn)
            config_layout.addRow(label_text, container)
            return line_edit

        self.input_format_path_tab2 = create_file_picker('format_path_tab2', "格式模板:")
        self.input_decorate_path_tab2 = create_file_picker('decorate_path_tab2', "样式模板:")

        # 参数配置
        param_layout = QHBoxLayout()

        self.input_stencilqty = QLineEdit(str(self.config_data['params']['StencilQty']))
        self.input_smtqty = QLineEdit(str(self.config_data['params']['SMTCarrierQty']))
        self.input_qty = QLineEdit(str(self.config_data['params']['Qty']))

        # 验证器：只允许输入数字
        for inp in [self.input_stencilqty, self.input_smtqty, self.input_qty]:
            inp.setFixedWidth(100)
            inp.setPlaceholderText("数字")

        param_layout.addWidget(QLabel("StencilQty:"))
        param_layout.addWidget(self.input_stencilqty)
        param_layout.addSpacing(20)
        param_layout.addWidget(QLabel("SMTCarrierQty:"))
        param_layout.addWidget(self.input_smtqty)
        param_layout.addSpacing(20)
        param_layout.addWidget(QLabel("Qty:"))
        param_layout.addWidget(self.input_qty)
        param_layout.addStretch()

        config_layout.addRow("参数设置:", param_layout)

        # 保存配置按钮
        save_cfg_btn = QPushButton("保存")
        save_cfg_btn.setFixedWidth(90)
        save_cfg_btn.clicked.connect(self.save_current_config_tab2)
        config_layout.addRow("", save_cfg_btn)

        config_group.setLayout(config_layout)
        tab2_layout.addWidget(config_group)

        # 3. 操作区域
        action_group = QGroupBox("一键报价")
        action_layout = QVBoxLayout()

        # 源文件选择
        source_layout = QHBoxLayout()
        self.source_file_edit_tab2 = QLineEdit()
        self.source_file_edit_tab2.setPlaceholderText("请选择要处理的文件夹...")
        self.source_file_edit_tab2.setReadOnly(True)

        select_source_btn = QPushButton("选择文件夹")
        select_source_btn.setObjectName("PrimaryButton")
        select_source_btn.setCursor(Qt.PointingHandCursor)
        select_source_btn.clicked.connect(self.select_source_file_tab2)

        source_layout.addWidget(self.source_file_edit_tab2)
        source_layout.addWidget(select_source_btn)

        # 运行按钮
        self.run_btn_tab2 = QPushButton("开始处理")
        self.run_btn_tab2.setObjectName("ActionButton")
        self.run_btn_tab2.setCursor(Qt.PointingHandCursor)
        self.run_btn_tab2.setFixedHeight(50)
        self.run_btn_tab2.setEnabled(False)  # 选择文件后启用
        self.run_btn_tab2.clicked.connect(self.start_processing_tab2)

        action_layout.addLayout(source_layout)
        action_layout.addSpacing(10)
        action_layout.addWidget(self.run_btn_tab2)
        action_group.setLayout(action_layout)
        tab2_layout.addWidget(action_group)

        # 4. 日志与进度
        log_group = QGroupBox("运行日志 (Logs)")
        log_layout = QVBoxLayout()

        self.log_area_tab2 = QTextEdit()
        self.log_area_tab2.setReadOnly(True)
        self.log_area_tab2.setObjectName("LogArea")

        self.progress_bar_tab2 = QProgressBar()
        self.progress_bar_tab2.setValue(0)
        self.progress_bar_tab2.setTextVisible(True)
        self.progress_bar_tab2.setAlignment(Qt.AlignCenter)

        log_layout.addWidget(self.log_area_tab2)
        log_layout.addWidget(self.progress_bar_tab2)
        log_group.setLayout(log_layout)

        tab2_layout.addWidget(log_group, stretch=1)
        
        tab2_layout.addStretch()

    def apply_stylesheet(self):
        # 现代商业风格 CSS
        style = """
        QMainWindow {
            background-color: #f5f6f8;
        }
        QLabel {
            color: #333333;
            font-family: 'Segoe UI', 'Microsoft YaHei';
            font-size: 14px;
        }
        QLabel#TitleLabel {
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
        }
        QGroupBox {
            background-color: white;
            border: 1px solid #dcdcdc;
            border-radius: 8px;
            margin-top: 10px;
            padding-top: 15px;
            font-weight: bold;
            color: #555;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top left;
            padding: 0 5px;
            left: 10px;
        }
        QLineEdit {
            border: 1px solid #ccc;
            border-radius: 4px;
            padding: 6px;
            background-color: #ffffff;
            selection-background-color: #3498db;
        }
        QLineEdit:focus {
            border: 1px solid #3498db;
        }
        QPushButton {
            background-color: #ecf0f1;
            border: 1px solid #bdc3c7;
            color: #333;
            border-radius: 4px;
            padding: 6px 15px;
            font-weight: 500;
        }
        QPushButton:hover {
            background-color: #bdc3c7;
        }
        QPushButton#PrimaryButton {
            background-color: #3498db;
            color: white;
            border: none;
        }
        QPushButton#PrimaryButton:hover {
            background-color: #2980b9;
        }
        QPushButton#ActionButton {
            background-color: #27ae60;
            color: white;
            font-size: 16px;
            border: none;
            border-radius: 6px;
        }
        QPushButton#ActionButton:hover {
            background-color: #2ecc71;
        }
        QPushButton#ActionButton:disabled {
            background-color: #95a5a6;
        }
        QTextEdit#LogArea {
            background-color: #2c3e50;
            color: #ecf0f1;
            border-radius: 4px;
            font-family: Consolas, Monospace;
        }
        QProgressBar {
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            text-align: center;
            background-color: #ecf0f1;
        }
        QProgressBar::chunk {
            background-color: #3498db;
        }
        """
        self.setStyleSheet(style)

    def browse_file(self, line_edit, key):
        path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel Files (*.xlsx)")
        if path:
            line_edit.setText(path)
            self.config_data['paths'][key] = path

    def select_source_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 Source 文件", "", "Excel Files (*.xlsx)")
        if path:
            self.source_file_edit.setText(path)
            self.run_btn.setEnabled(True)
            self.log_message(f"已加载源文件: {path}")

    def select_source_file_tab2(self):
        path = QFileDialog.getExistingDirectory(self, "选择源文件夹")
        if path:
            self.source_file_edit_tab2.setText(path)
            self.run_btn_tab2.setEnabled(True)
            self.log_message_tab2(f"已加载源文件夹: {path}")

    def save_current_config(self):
        # 更新配置对象
        self.config_data['paths']['format_path'] = self.input_format_path.text()
        self.config_data['paths']['decorate_path'] = self.input_decorate_path.text()

        try:
            self.config_data['params']['Handling'] = float(self.input_handling.text())
            self.config_data['params']['UPUSDL'] = float(self.input_upusdl.text())
            self.config_data['params']['UPUSOH'] = float(self.input_upusoh.text())
            self.config_data['params']['StencilQty'] = float(self.input_stencilqty.text())
            self.config_data['params']['SMTCarrierQty'] = float(self.input_smtqty.text())
            self.config_data['params']['Qty'] = float(self.input_qty.text())

            ConfigManager.save_config(self.config_data)
            QMessageBox.information(self, "成功", "配置已保存！")
            self.log_message("配置已更新并保存。")
        except ValueError:
            QMessageBox.warning(self, "错误", "参数必须是有效的数字！")

    def save_current_config_tab2(self):
        # 更新配置对象
        self.config_data['paths']['format_path_tab2'] = self.input_format_path_tab2.text()
        self.config_data['paths']['decorate_path_tab2'] = self.input_decorate_path_tab2.text()

        try:
            self.config_data['params']['StencilQty'] = float(self.input_stencilqty.text())
            self.config_data['params']['SMTCarrierQty'] = float(self.input_smtqty.text())
            self.config_data['params']['Qty'] = float(self.input_qty.text())

            ConfigManager.save_config(self.config_data)
            QMessageBox.information(self, "成功", "配置已保存！")
            self.log_message_tab2("配置已更新并保存。")
        except ValueError:
            QMessageBox.warning(self, "错误", "参数必须是有效的数字！")

    def log_message(self, msg):
        timestamp = datetime.now().strftime("[%H:%M:%S] ")
        self.log_area.append(timestamp + msg)
        # 滚动到底部
        self.log_area.verticalScrollBar().setValue(self.log_area.verticalScrollBar().maximum())

    def log_message_tab2(self, msg):
        timestamp = datetime.now().strftime("[%H:%M:%S] ")
        self.log_area_tab2.append(timestamp + msg)
        # 滚动到底部
        self.log_area_tab2.verticalScrollBar().setValue(self.log_area_tab2.verticalScrollBar().maximum())

    def start_processing(self):
        source_path = self.source_file_edit.text()
        if not source_path:
            return

        # 锁定界面
        self.run_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_message("任务开始...")

        # 确保当前界面的参数也传进去
        current_config = copy(self.config_data)
        try:
            current_config['params']['Handling'] = float(self.input_handling.text())
            current_config['params']['UPUSDL'] = float(self.input_upusdl.text())
            current_config['params']['UPUSOH'] = float(self.input_upusoh.text())
            current_config['params']['StencilQty'] = float(self.input_stencilqty.text())
            current_config['params']['SMTCarrierQty'] = float(self.input_smtqty.text())
            current_config['params']['Qty'] = float(self.input_qty.text())
        except ValueError:
            QMessageBox.critical(self, "错误", "参数格式不正确")
            self.run_btn.setEnabled(True)
            return

        self.worker = ExcelWorker(source_path, current_config)
        self.worker.log_signal.connect(self.log_message)
        self.worker.progress_signal.connect(self.progress_bar.setValue)
        self.worker.finished_signal.connect(self.on_processing_finished)
        self.worker.start()

    def start_processing_tab2(self):
        source_path = self.source_file_edit_tab2.text()
        if not source_path:
            return

        # 锁定界面
        self.run_btn_tab2.setEnabled(False)
        self.progress_bar_tab2.setValue(0)
        self.log_message_tab2("任务开始...")

        # 确保当前界面的参数也传进去
        current_config = copy(self.config_data)
        try:
            current_config['params']['StencilQty'] = float(self.input_stencilqty.text())
            current_config['params']['SMTCarrierQty'] = float(self.input_smtqty.text())
            current_config['params']['Qty'] = float(self.input_qty.text())
        except ValueError:
            QMessageBox.critical(self, "错误", "参数格式不正确")
            self.run_btn_tab2.setEnabled(True)
            return

        self.worker2 = ExcelWorker2(source_path, current_config)
        self.worker2.log_signal.connect(self.log_message_tab2)
        self.worker2.progress_signal.connect(self.progress_bar_tab2.setValue)
        self.worker2.finished_signal.connect(self.on_processing_finished_tab2)
        self.worker2.start()

    def on_processing_finished(self, success, message):
        self.run_btn.setEnabled(True)
        self.progress_bar.setValue(100 if success else 0)
        self.log_message("任务结束。")

        if success:
            QMessageBox.information(self, "处理完成", message)
        else:
            QMessageBox.critical(self, "处理失败", message)

    def on_processing_finished_tab2(self, success, message):
        self.run_btn_tab2.setEnabled(True)
        self.progress_bar_tab2.setValue(100 if success else 0)
        self.log_message_tab2("任务结束。")

        if success:
            QMessageBox.information(self, "处理完成", message)
        else:
            QMessageBox.critical(self, "处理失败", message)
    
    def example_button_clicked(self):
        QMessageBox.information(self, "提示", "这是第二个选项卡的示例功能！")


# =============================================================================
# 5. 程序入口
# =============================================================================
if __name__ == "__main__":
    # 创建必要的文件夹结构，防止报错
    if not os.path.exists("LogicData"):
        os.makedirs("LogicData")
        print("提示: 请确保 LogicData 文件夹下存在 Format.xlsx 和 Decorate.xlsx")

    app = QApplication(sys.argv)

    # 设置应用字体
    font = QFont("Microsoft YaHei", 9)
    app.setFont(font)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())